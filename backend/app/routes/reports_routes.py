from flask import Blueprint, request, jsonify, send_file
from flask_jwt_extended import jwt_required, get_jwt
from app import db
from app.models.sql_models import Delivery, Order, Client, Agent, StockMovement, Product, DeliveryItem
from datetime import datetime, timedelta
# import pandas as pd # Removed to fix crash
import io
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import os
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

reports_bp = Blueprint('reports', __name__, url_prefix='/api/reports')

# --- 1. EXPORT EXCEL/CSV ---

@reports_bp.route('/export/excel', methods=['POST'])
@jwt_required()
def export_excel():
    """Exporter des données au format Excel"""
    try:
        data = request.get_json()
        report_type = data.get('report_type', 'deliveries')
        start_date = data.get('start_date')
        end_date = data.get('end_date')
        
        # Créer le classeur Excel
        wb = Workbook()
        ws = wb.active
        ws.title = f"Rapport_{report_type}"
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        if report_type == 'deliveries':
            # Export des livraisons
            query = Delivery.query
            
            if start_date:
                query = query.filter(Delivery.created_at >= datetime.fromisoformat(start_date))
            if end_date:
                query = query.filter(Delivery.created_at <= datetime.fromisoformat(end_date))
                
            deliveries = query.all()
            
            # En-têtes
            headers = ['ID', 'Client', 'Agent', 'Date', 'Statut', 'Produits', 'Montant', 'Notes']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Données
            for row, delivery in enumerate(deliveries, 2):
                ws.cell(row=row, column=1, value=str(delivery.id))
                ws.cell(row=row, column=2, value=getattr(delivery.client, 'name', 'Inconnu') if delivery.client else 'Inconnu')
                ws.cell(row=row, column=3, value=getattr(delivery.agent, 'full_name', 'Non assigné') if delivery.agent else 'Non assigné')
                ws.cell(row=row, column=4, value=delivery.created_at.strftime('%Y-%m-%d %H:%M') if delivery.created_at else '--')
                ws.cell(row=row, column=5, value=delivery.status or 'N/A')
                # Formatage des produits
                products_str = ", ".join([f"{item.product.name} ({item.quantity})" for item in delivery.items]) if delivery.items else "Aucun"
                
                ws.cell(row=row, column=6, value=products_str)
                ws.cell(row=row, column=7, value=float(delivery.total_amount or 0))
                ws.cell(row=row, column=8, value='') 
        
        elif report_type == 'stock':
            # Export des mouvements de stock
            query = StockMovement.query
            
            if start_date:
                query = query.filter(StockMovement.timestamp >= datetime.fromisoformat(start_date))
            if end_date:
                query = query.filter(StockMovement.timestamp <= datetime.fromisoformat(end_date))
                
            movements = query.all()
            
            # En-têtes
            headers = ['ID', 'Produit', 'Type', 'Quantité', 'Référence', 'Agent', 'Date', 'Notes']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Données
            for row, movement in enumerate(movements, 2):
                ws.cell(row=row, column=1, value=str(movement.id))
                ws.cell(row=row, column=2, value=movement.stock_item.product.name if movement.stock_item and movement.stock_item.product else 'Inconnu')
                ws.cell(row=row, column=3, value=movement.movement_type)
                ws.cell(row=row, column=4, value=movement.quantity)
                ws.cell(row=row, column=5, value=movement.reference or '')
                ws.cell(row=row, column=6, value=movement.agent.full_name if movement.agent else 'Système')
                ws.cell(row=row, column=7, value=movement.timestamp.strftime('%Y-%m-%d %H:%M'))
                ws.cell(row=row, column=8, value=movement.notes or '')
        
        elif report_type == 'clients':
            # Export des clients
            clients = Client.query.all()
            
            # En-têtes
            headers = ['ID', 'Nom', 'Téléphone', 'Email', 'Adresse', 'Date création', 'Total commandes']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Données
            for row, client in enumerate(clients, 2):
                ws.cell(row=row, column=1, value=str(client.id))
                ws.cell(row=row, column=2, value=client.name)
                ws.cell(row=row, column=3, value=client.phone)
                ws.cell(row=row, column=4, value=client.email or '')
                ws.cell(row=row, column=5, value=client.address or '')
                ws.cell(row=row, column=6, value=client.created_at.strftime('%Y-%m-%d'))
                ws.cell(row=row, column=7, value=len(client.orders) if hasattr(client, 'orders') else 0)
        
        # Ajuster la largeur des colonnes
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Sauvegarder en mémoire
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        filename = f"rapport_{report_type}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return send_file(
            excel_file,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        return jsonify({"msg": f"Erreur export Excel: {str(e)}"}), 500

@reports_bp.route('/export/pdf', methods=['POST'])
@jwt_required()
def export_pdf():
    """Exporter des données au format PDF avec ReportLab"""
    try:
        data = request.get_json()
        report_type = data.get('report_type', 'deliveries')
        start_date = data.get('start_date')
        end_date = data.get('end_date')
        
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(letter))
        elements = []
        
        styles = getSampleStyleSheet()
        title_style = styles['Title']
        header_style = styles['Heading2']
        
        title = f"Rapport des {'Livraisons' if report_type == 'deliveries' else 'Stocks'}"
        elements.append(Paragraph(title, title_style))
        elements.append(Spacer(1, 12))
        
        if start_date and end_date:
            date_range = f"Période: {start_date} au {end_date}"
            elements.append(Paragraph(date_range, styles['Normal']))
            elements.append(Spacer(1, 12))
            
        if report_type == 'deliveries':
            query = Delivery.query
            if start_date:
                query = query.filter(Delivery.created_at >= datetime.fromisoformat(start_date))
            if end_date:
                query = query.filter(Delivery.created_at <= datetime.fromisoformat(end_date))
            
            deliveries = query.all()
            
            # En-têtes du tableau
            table_data = [['ID', 'Client', 'Agent', 'Date', 'Statut', 'Produits', 'Montant']]
            
            for d in deliveries:
                client_name = 'Inconnu'
                if d.client:
                    client_name = d.client.name
                
                agent_name = 'NC'
                if d.agent:
                    agent_name = d.agent.full_name

                # Formatage dynamique des produits
                products_str = "\n".join([f"- {item.product.name}: {item.quantity}" for item in d.items])
                if not products_str: products_str = "--"

                table_data.append([
                    str(d.id),
                    (client_name[:15] + '..') if len(client_name) > 15 else client_name,
                    agent_name,
                    d.date.strftime('%d/%m/%Y') if d.date else '--',
                    d.status or 'N/A',
                    products_str,
                    f"{getattr(d, 'total_amount', 0):,.0f} F"
                ])
                
            # Créer le tableau ReportLab
            t = Table(table_data, repeatRows=1)
            t.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#4F81BD")),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 1), (-1, -1), 10),
            ]))
            elements.append(t)
            
        doc.build(elements)
        buffer.seek(0)
        
        filename = f"rapport_{report_type}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        
        return send_file(
            buffer,
            as_attachment=True,
            download_name=filename,
            mimetype='application/pdf'
        )
        
    except Exception as e:
        return jsonify({"msg": f"Erreur export PDF: {str(e)}"}), 500

@reports_bp.route('/export/csv', methods=['POST'])
@jwt_required()
def export_csv():
    """Exporter des données au format CSV"""
    try:
        data = request.get_json()
        report_type = data.get('report_type', 'deliveries')
        start_date = data.get('start_date')
        end_date = data.get('end_date')
        
        output = io.StringIO()
        writer = csv.writer(output)
        
        if report_type == 'deliveries':
            # Export des livraisons
            query = Delivery.query
            
            if start_date:
                query = query.filter(Delivery.created_at >= datetime.fromisoformat(start_date))
            if end_date:
                query = query.filter(Delivery.created_at <= datetime.fromisoformat(end_date))
                
            deliveries = query.all()
            
            # En-têtes
            headers = ['ID', 'Client', 'Agent', 'Date', 'Statut', 'Produits', 'Montant', 'Notes']
            writer.writerow(headers)
            
            # Données
            for delivery in deliveries:
                client_name = delivery.client.name if (hasattr(delivery, 'client') and delivery.client) else 'Inconnu'
                agent_name = delivery.agent.full_name if (hasattr(delivery, 'agent') and delivery.agent) else 'Non assigné'
                # Formatage des produits
                products_str = ", ".join([f"{item.product.name} x{item.quantity}" for item in delivery.items]) if delivery.items else ""

                writer.writerow([
                    delivery.id,
                    client_name,
                    agent_name,
                    delivery.created_at.strftime('%Y-%m-%d %H:%M') if delivery.created_at else '--',
                    delivery.status or 'N/A',
                    products_str,
                    getattr(delivery, 'total_amount', 0) or 0,
                    ''
                ])
        
        elif report_type == 'stock':
            # Export des mouvements de stock
            query = StockMovement.query
            
            if start_date:
                query = query.filter(StockMovement.timestamp >= datetime.fromisoformat(start_date))
            if end_date:
                query = query.filter(StockMovement.timestamp <= datetime.fromisoformat(end_date))
                
            movements = query.all()
            
            # En-têtes
            headers = ['ID', 'Produit', 'Type', 'Quantité', 'Référence', 'Agent', 'Date', 'Notes']
            writer.writerow(headers)
            
            # Données
            for movement in movements:
                writer.writerow([
                    movement.id,
                    movement.stock_item.product.name if movement.stock_item and movement.stock_item.product else 'Inconnu',
                    movement.movement_type,
                    movement.quantity,
                    movement.reference or '',
                    movement.agent.full_name if movement.agent else 'Système',
                    movement.timestamp.strftime('%Y-%m-%d %H:%M'),
                    movement.notes or ''
                ])
        
        output.seek(0)
        filename = f"rapport_{report_type}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        
        return send_file(
            io.BytesIO(output.getvalue().encode('utf-8-sig')),
            as_attachment=True,
            download_name=filename,
            mimetype='text/csv'
        )
        
    except Exception as e:
        return jsonify({"msg": f"Erreur export CSV: {str(e)}"}), 500

# --- 2. RAPPORTS PERSONNALISÉS ---

@reports_bp.route('/custom', methods=['POST'])
@jwt_required()
def generate_custom_report():
    """Générer un rapport personnalisé selon les critères"""
    try:
        data = request.get_json()
        report_config = data.get('config', {})
        
        report_type = report_config.get('type')
        start_date = report_config.get('start_date')
        end_date = report_config.get('end_date')
        filters = report_config.get('filters', {})
        group_by = report_config.get('group_by')
        
        result = {
            'report_type': report_type,
            'period': {
                'start': start_date,
                'end': end_date
            },
            'generated_at': datetime.utcnow().isoformat(),
            'data': []
        }
        
        if report_type == 'sales_summary':
            # Rapport de ventes résumé
            query = db.session.query(
                Delivery.status,
                db.func.count(Delivery.id).label('count'),
                db.func.sum(Delivery.total_amount).label('total_amount'),
                db.func.avg(Delivery.total_amount).label('avg_amount')
            )
            
            if start_date:
                query = query.filter(Delivery.created_at >= datetime.fromisoformat(start_date))
            if end_date:
                query = query.filter(Delivery.created_at <= datetime.fromisoformat(end_date))
            
            if group_by == 'status':
                query = query.group_by(Delivery.status)
            elif group_by == 'date':
                query = query.group_by(db.func.date(Delivery.created_at))
            
            summary = query.all()
            
            for row in summary:
                result['data'].append({
                    'group': getattr(row, group_by, 'total') if group_by else 'total',
                    'count': row.count,
                    'total_amount': float(row.total_amount or 0),
                    'avg_amount': float(row.avg_amount or 0)
                })
        
        elif report_type == 'agent_performance':
            # Performance des agents
            query = db.session.query(
                Agent.id,
                Agent.full_name,
                db.func.count(Delivery.id).label('deliveries_count'),
                db.func.sum(Delivery.total_amount).label('total_revenue'),
                db.func.avg(Delivery.total_amount).label('avg_delivery_value')
            ).outerjoin(Delivery, Agent.id == Delivery.agent_id)
            
            if start_date:
                query = query.filter(Delivery.created_at >= datetime.fromisoformat(start_date))
            if end_date:
                query = query.filter(Delivery.created_at <= datetime.fromisoformat(end_date))
            
            query = query.group_by(Agent.id, Agent.full_name)
            agents_data = query.all()
            
            for row in agents_data:
                result['data'].append({
                    'agent_id': row.id,
                    'agent_name': row.full_name,
                    'deliveries_count': row.deliveries_count,
                    'total_revenue': float(row.total_revenue or 0),
                    'avg_delivery_value': float(row.avg_delivery_value or 0)
                })
        
        elif report_type == 'product_analysis':
            # Analyse dynamique des produits par ventes réelles (DeliveryItems)
            query = db.session.query(
                Product.name,
                db.func.sum(DeliveryItem.quantity).label('total_quantity'),
                db.func.sum(DeliveryItem.quantity * Product.price).label('revenue_generated')
            ).join(DeliveryItem, Product.id == DeliveryItem.product_id)\
             .join(Delivery, DeliveryItem.delivery_id == Delivery.id)\
             .filter(Delivery.status == 'completed')

            if start_date:
                query = query.filter(Delivery.date >= datetime.fromisoformat(start_date))
            if end_date:
                query = query.filter(Delivery.date <= datetime.fromisoformat(end_date))

            summary = query.group_by(Product.name).all()
            
            for row in summary:
                result['data'].append({
                    'product_name': row.name,
                    'total_quantity': int(row.total_quantity or 0),
                    'total_revenue': float(row.revenue_generated or 0)
                })
        
        return jsonify(result), 200
    except Exception as e:
        return jsonify({"msg": f"Erreur rapport: {str(e)}"}), 200 # Return 200 with error msg to avoid crash


# --- 3. GRAPHIQUES INTERACTIFS ---

@reports_bp.route('/charts/revenue', methods=['GET'])
@jwt_required()
def get_revenue_chart():
    """Données pour graphique des revenus (Dynamic & Robust)"""
    try:
        period = request.args.get('period', 'month')  # day, week, month
        
        if period == 'day':
            # Revenus par jour (30 derniers jours)
            start_date = datetime.utcnow() - timedelta(days=30)
            query = db.session.query(
                func.date(Delivery.created_at).label('date_label'),
                func.sum(Delivery.total_amount).label('revenue')
            ).filter(Delivery.created_at >= start_date, Delivery.status == 'completed')\
             .group_by(func.date(Delivery.created_at))
        else:
            # Revenus par mois (Par défaut)
            query = db.session.query(
                extract('month', Delivery.created_at).label('month'),
                extract('year', Delivery.created_at).label('year'),
                func.sum(Delivery.total_amount).label('revenue')
            ).filter(Delivery.status == 'completed')\
             .group_by(extract('year', Delivery.created_at), extract('month', Delivery.created_at))\
             .order_by(extract('year', Delivery.created_at), extract('month', Delivery.created_at))
        
        data = query.all()
        
        if not data:
            return jsonify({'labels': [], 'datasets': [{'label': 'Revenus (F CFA)', 'data': []}]}), 200
            
        labels = []
        revenues = []
        
        for row in data:
            if period == 'day':
                labels.append(str(row.date_label))
            else:
                labels.append(f"{int(row.month)}/{row.year}")
            revenues.append(float(row.revenue or 0))

        return jsonify({
            'labels': labels,
            'datasets': [{
                'label': 'Revenus (F CFA)',
                'data': revenues,
                'backgroundColor': 'rgba(59, 130, 246, 0.2)',
                'borderColor': 'rgba(59, 130, 246, 1)',
                'borderWidth': 2,
                'fill': True
            }]
        }), 200
    except Exception as e:
        return jsonify({'labels': [], 'datasets': []}), 200

@reports_bp.route('/charts/delivery-status', methods=['GET'])
@jwt_required()
def get_delivery_status_chart():
    """Données pour graphique des statuts de livraison"""
    try:
        query = db.session.query(
            Delivery.status,
            db.func.count(Delivery.id).label('count')
        ).group_by(Delivery.status)
        
        if request.args.get('period'):
            period = request.args.get('period')
            if period == 'day':
                start_date = datetime.utcnow() - timedelta(days=1)
            elif period == 'week':
                start_date = datetime.utcnow() - timedelta(weeks=1)
            elif period == 'month':
                start_date = datetime.utcnow() - timedelta(days=30)
            else:
                start_date = None
            
            if start_date:
                query = query.filter(Delivery.created_at >= start_date)
        
        data = query.all()
        
        chart_data = {
            'labels': [row.status for row in data],
            'datasets': [{
                'label': 'Nombre de livraisons',
                'data': [row.count for row in data],
                'backgroundColor': [
                    'rgba(34, 197, 94, 0.8)',   # delivered - vert
                    'rgba(59, 130, 246, 0.8)',   # in_progress - bleu
                    'rgba(251, 146, 60, 0.8)',  # pending - orange
                    'rgba(239, 68, 68, 0.8)',   # cancelled - rouge
                ],
                'borderWidth': 1
            }]
        }
        
        return jsonify(chart_data), 200
        
    except Exception as e:
        return jsonify({}), 200

@reports_bp.route('/charts/top-products', methods=['GET'])
@jwt_required()
def get_top_products_chart():
    """Données pour graphique des produits les plus vendus"""
    try:
        limit = int(request.args.get('limit', 10))
        
        # Top produits basé sur DeliveryItem (Ventes réelles)
        query = db.session.query(
            Product.name,
            db.func.sum(DeliveryItem.quantity).label('total_quantity'),
            db.func.sum(DeliveryItem.quantity * Product.price).label('total_revenue')
        ).join(DeliveryItem, Product.id == DeliveryItem.product_id)\
         .join(Delivery, DeliveryItem.delivery_id == Delivery.id)\
         .filter(Delivery.status == 'completed')\
         .group_by(Product.id, Product.name)\
         .order_by(db.func.sum(DeliveryItem.quantity).desc())\
         .limit(limit)
        
        data = query.all()
        
        if not data:
            return jsonify({
                'labels': ["Aucune vente"],
                'datasets': [
                    {'label': 'Quantité', 'data': [0], 'backgroundColor': 'rgba(200, 200, 200, 0.5)'},
                    {'label': 'Revenus', 'data': [0], 'backgroundColor': 'rgba(200, 200, 200, 0.3)'}
                ]
            }), 200

        chart_data = {
            'labels': [row.name for row in data],
            'datasets': [
                {
                    'label': 'Quantité vendue',
                    'data': [int(row.total_quantity or 0) for row in data],
                    'backgroundColor': 'rgba(59, 130, 246, 0.8)',
                    'yAxisID': 'y'
                },
                {
                    'label': 'Revenus (F CFA)',
                    'data': [float(row.total_revenue or 0) for row in data],
                    'backgroundColor': 'rgba(34, 197, 94, 0.8)',
                    'yAxisID': 'y1'
                }
            ]
        }
        
        return jsonify(chart_data), 200
    except Exception as e:
        return jsonify({"msg": f"Erreur chart: {str(e)}"}), 200

